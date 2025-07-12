from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, 
                             QMessageBox, QLabel, QComboBox, QFrame, QGridLayout, QFileDialog, QProgressBar)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
import csv
import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from database.init_db import get_connection

class ExportWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)
    
    def __init__(self, export_type, format_type, filename):
        super().__init__()
        self.export_type = export_type
        self.format_type = format_type
        self.filename = filename
    
    def run(self):
        try:
            if self.format_type == "CSV":
                self.export_csv()
            elif self.format_type == "PDF":
                self.export_pdf()
            elif self.format_type == "Excel":
                self.export_excel()
            self.finished.emit(f"Export completed: {self.filename}")
        except Exception as e:
            self.error.emit(f"Export failed: {str(e)}")
    
    def export_csv(self):
        conn = get_connection()
        c = conn.cursor()
        
        if self.export_type == "All Data":
            # Export all tables
            tables = ['batches', 'feed_logs', 'water_logs', 'vaccinations', 'mortality', 'workers', 'expenses', 'revenue']
            with open(self.filename, 'w', newline='', encoding='utf-8') as file:
                for i, table in enumerate(tables):
                    c.execute(f'SELECT * FROM {table}')
                    rows = c.fetchall()
                    if rows:
                        # Get column names
                        c.execute(f'PRAGMA table_info({table})')
                        columns = [col[1] for col in c.fetchall()]
                        
                        writer = csv.writer(file)
                        writer.writerow([f'=== {table.upper()} ==='])
                        writer.writerow(columns)
                        writer.writerows(rows)
                        writer.writerow([])  # Empty row between tables
                    
                    self.progress.emit((i + 1) * 100 // len(tables))
        
        else:
            # Export specific table
            table_map = {
                "Batches": ("batches", ["Batch ID", "Num Chicks", "Breed", "Date In", "Expected Out", "Mortality Rate"]),
                "Feed/Water Logs": ("feed_logs", ["Batch ID", "Date", "Quantity (kg)"]),
                "Vaccinations": ("vaccinations", ["Batch ID", "Date", "Vaccine", "Status"]),
                "Mortality": ("mortality", ["Batch ID", "Date", "Count", "Reason"]),
                "Workers": ("workers", ["Worker ID", "Name", "Role", "Phone", "Email", "Address", "Salary", "Hire Date", "Status"]),
                "Expenses": ("expenses", ["Date", "Category", "Amount", "Description", "Payment Method"]),
                "Revenue": ("revenue", ["Date", "Batch ID", "Amount"])
            }
            
            table_name, headers = table_map[self.export_type]
            c.execute(f'SELECT * FROM {table_name}')
            rows = c.fetchall()
            
            with open(self.filename, 'w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(headers)
                writer.writerows(rows)
        
        conn.close()
    
    def export_pdf(self):
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            doc = SimpleDocTemplate(self.filename, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph("Dash Poultry - Data Export", title_style))
            story.append(Spacer(1, 20))
            
            conn = get_connection()
            c = conn.cursor()
            
            if self.export_type == "All Data":
                tables = ['batches', 'feed_logs', 'water_logs', 'vaccinations', 'mortality', 'workers', 'expenses', 'revenue']
                for table in tables:
                    c.execute(f'SELECT * FROM {table}')
                    rows = c.fetchall()
                    if rows:
                        # Get column names
                        c.execute(f'PRAGMA table_info({table})')
                        columns = [col[1] for col in c.fetchall()]
                        
                        # Create table
                        data = [columns] + [list(row) for row in rows]
                        table_obj = Table(data)
                        table_obj.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 12),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        
                        story.append(Paragraph(f"{table.upper()}", styles['Heading2']))
                        story.append(table_obj)
                        story.append(Spacer(1, 20))
            else:
                # Export specific table
                table_map = {
                    "Batches": ("batches", ["Batch ID", "Num Chicks", "Breed", "Date In", "Expected Out", "Mortality Rate"]),
                    "Feed/Water Logs": ("feed_logs", ["Batch ID", "Date", "Quantity (kg)"]),
                    "Vaccinations": ("vaccinations", ["Batch ID", "Date", "Vaccine", "Status"]),
                    "Mortality": ("mortality", ["Batch ID", "Date", "Count", "Reason"]),
                    "Workers": ("workers", ["Worker ID", "Name", "Role", "Phone", "Email", "Address", "Salary", "Hire Date", "Status"]),
                    "Expenses": ("expenses", ["Date", "Category", "Amount", "Description", "Payment Method"]),
                    "Revenue": ("revenue", ["Date", "Batch ID", "Amount"])
                }
                
                table_name, headers = table_map[self.export_type]
                c.execute(f'SELECT * FROM {table_name}')
                rows = c.fetchall()
                
                data = [headers] + [list(row) for row in rows]
                table_obj = Table(data)
                table_obj.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(table_obj)
            
            conn.close()
            doc.build(story)
            
        except ImportError:
            raise Exception("PDF export requires reportlab library. Install with: pip install reportlab")
    
    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            conn = get_connection()
            c = conn.cursor()
            
            if self.export_type == "All Data":
                tables = ['batches', 'feed_logs', 'water_logs', 'vaccinations', 'mortality', 'workers', 'expenses', 'revenue']
                for i, table in enumerate(tables):
                    if i == 0:
                        ws = wb.active
                        ws.title = table.capitalize()
                    else:
                        ws = wb.create_sheet(table.capitalize())
                    
                    c.execute(f'SELECT * FROM {table}')
                    rows = c.fetchall()
                    if rows:
                        # Get column names
                        c.execute(f'PRAGMA table_info({table})')
                        columns = [col[1] for col in c.fetchall()]
                        
                        # Write headers
                        for col, header in enumerate(columns, 1):
                            cell = ws.cell(row=1, column=col, value=header)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        
                        # Write data
                        for row_idx, row in enumerate(rows, 2):
                            for col_idx, value in enumerate(row, 1):
                                ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    self.progress.emit((i + 1) * 100 // len(tables))
            else:
                # Export specific table
                table_map = {
                    "Batches": ("batches", ["Batch ID", "Num Chicks", "Breed", "Date In", "Expected Out", "Mortality Rate"]),
                    "Feed/Water Logs": ("feed_logs", ["Batch ID", "Date", "Quantity (kg)"]),
                    "Vaccinations": ("vaccinations", ["Batch ID", "Date", "Vaccine", "Status"]),
                    "Mortality": ("mortality", ["Batch ID", "Date", "Count", "Reason"]),
                    "Workers": ("workers", ["Worker ID", "Name", "Role", "Phone", "Email", "Address", "Salary", "Hire Date", "Status"]),
                    "Expenses": ("expenses", ["Date", "Category", "Amount", "Description", "Payment Method"]),
                    "Revenue": ("revenue", ["Date", "Batch ID", "Amount"])
                }
                
                table_name, headers = table_map[self.export_type]
                c.execute(f'SELECT * FROM {table_name}')
                rows = c.fetchall()
                
                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Write data
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            conn.close()
            wb.save(self.filename)
            
        except ImportError:
            raise Exception("Excel export requires openpyxl library. Install with: pip install openpyxl")

class ExportModuleWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.export_worker = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Title
        title = QLabel("<b>Data Export</b>")
        title.setStyleSheet("font-size: 18px; margin-bottom: 8px;")
        layout.addWidget(title)
        
        # Export options
        options_frame = QFrame()
        options_frame.setFrameStyle(QFrame.Shape.Box)
        options_frame.setStyleSheet("QFrame { background-color: #f9fafb; border: 1px solid #e5e7eb; border-radius: 8px; padding: 10px; }")
        options_layout = QVBoxLayout(options_frame)
        
        # Data selection
        data_layout = QHBoxLayout()
        data_layout.addWidget(QLabel("Export Data:"))
        self.data_combo = QComboBox()
        self.data_combo.addItems(["All Data", "Batches", "Feed/Water Logs", "Vaccinations", "Mortality", "Workers", "Expenses", "Revenue"])
        data_layout.addWidget(self.data_combo)
        data_layout.addStretch()
        options_layout.addLayout(data_layout)
        
        # Format selection
        format_layout = QHBoxLayout()
        format_layout.addWidget(QLabel("Export Format:"))
        self.format_combo = QComboBox()
        self.format_combo.addItems(["CSV", "PDF", "Excel"])
        format_layout.addWidget(self.format_combo)
        format_layout.addStretch()
        options_layout.addLayout(format_layout)
        
        layout.addWidget(options_frame)
        
        # Export button
        self.export_btn = QPushButton("Export Data")
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #059669;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #047857;
            }
            QPushButton:disabled {
                background-color: #9ca3af;
            }
        """)
        self.export_btn.clicked.connect(self.start_export)
        layout.addWidget(self.export_btn)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Status label
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #6b7280; font-style: italic;")
        layout.addWidget(self.status_label)
        
        layout.addStretch()

    def start_export(self):
        export_type = self.data_combo.currentText()
        format_type = self.format_combo.currentText()
        
        # Get filename
        if format_type == "CSV":
            filename, _ = QFileDialog.getSaveFileName(self, "Export Data", "", "CSV Files (*.csv)")
        elif format_type == "PDF":
            filename, _ = QFileDialog.getSaveFileName(self, "Export Data", "", "PDF Files (*.pdf)")
        elif format_type == "Excel":
            filename, _ = QFileDialog.getSaveFileName(self, "Export Data", "", "Excel Files (*.xlsx)")
        
        if filename:
            self.export_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.status_label.setText("Exporting data...")
            
            self.export_worker = ExportWorker(export_type, format_type, filename)
            self.export_worker.progress.connect(self.progress_bar.setValue)
            self.export_worker.finished.connect(self.export_finished)
            self.export_worker.error.connect(self.export_error)
            self.export_worker.start()

    def export_finished(self, message):
        self.export_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.status_label.setText(message)
        QMessageBox.information(self, "Export Complete", message)

    def export_error(self, error_message):
        self.export_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.status_label.setText("Export failed")
        QMessageBox.critical(self, "Export Error", error_message) 